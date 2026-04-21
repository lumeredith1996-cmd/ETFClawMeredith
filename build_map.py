#!/usr/bin/env python3
"""
全球ETF投资地图 - 生成脚本
从 Excel 读取产品清单，生成完整的 index.html
"""
import openpyxl, json, os
from datetime import datetime

TODAY = datetime.now().strftime('%Y-%m-%d')
EXCEL_FILE = '/root/.openclaw/media/inbound/all产品_20260421.xlsx'
OUTPUT_FILE = '/root/.openclaw/workspace/etf-map-site/index.html'

# ── 读取产品清单 ─────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb['products_all']
products = {}
for row in ws.iter_rows(values_only=True):
    if row[0] is None or row[0] == '一级分类': continue
    code = str(row[2]).strip() if row[2] else ''
    if not code: continue
    products[code] = {
        'cat1': str(row[0]).strip(),
        'cat2': str(row[1]).strip(),
        'name': str(row[3]).strip() if row[3] else '',
        'note': str(row[4]).strip() if row[4] else '',
    }

mainland_prods = {k: v for k, v in products.items() if v['cat1'] == '内地可买'}
overseas_prods = {k: v for k, v in products.items() if v['cat1'] == '境外可买'}
print(f"内陆: {len(mainland_prods)} 境外: {len(overseas_prods)}")

# ── 持仓数据 ─────────────────────────────────────────────────────────────────
HOLDINGS_DIR = '/root/.openclaw/workspace/data'
holdings_cache = {}

def holdings_html_from_excel(code):
    fpath = os.path.join(HOLDINGS_DIR, f'{code}.xlsx')
    if not os.path.exists(fpath): return None
    try:
        wb2 = openpyxl.load_workbook(fpath)
        ws2 = wb2.active
        rows = ''
        for row in ws2.iter_rows(values_only=True):
            if row[0] is None or str(row[0]).startswith('持仓'): continue
            name, scode, region, w1, m1, ytd, wgt = [str(x).strip() if x else '' for x in row[:7]]
            if not name: continue
            def cc(v):
                if v.startswith('+'): return '#10b981'
                if v.startswith('-'): return '#ef4444'
                return '#9ca3af'
            def wc(v):
                try: return '#ef4444' if float(v.rstrip('%')) >= 5 else '#10b981'
                except: return '#10b981'
            rows += f"<tr><td style='padding:5px 8px;font-size:12px'>{name}</td><td style='padding:5px 8px;font-size:11px;color:#3b82f6;font-family:monospace'>{scode}</td><td style='padding:5px 8px;font-size:11px'>{region}</td><td style='padding:5px 8px;font-size:11px;text-align:right;color:{cc(w1)}'>{w1}</td><td style='padding:5px 8px;font-size:11px;text-align:right;color:{cc(m1)}'>{m1}</td><td style='padding:5px 8px;font-size:11px;text-align:right;color:{cc(ytd)}'>{ytd}</td><td style='padding:5px 8px;font-size:11px;text-align:right;font-weight:700;color:{wc(wgt)}'>{wgt}</td></tr>"
        return rows
    except Exception as e:
        print(f"  {code} holdings error: {e}")
        return None

for code in products:
    rows = holdings_html_from_excel(code)
    if rows:
        holdings_cache[code] = rows

# 3431.HK 硬编码数据（图片识别 2026-04-10）
H3431 = [
    ("美团-W","3690.HK","中国","+9.16%","+10.75%","-14.79%","7.50%"),
    ("阿里巴巴","9988.HK","中国","+5.91%","-5.99%","-12.11%","6.16%"),
    ("中芯国际","0981.HK","中国","+14.22%","-9.27%","-18.36%","6.09%"),
    ("小米集团","1810.HK","中国","+0.06%","-8.09%","-21.17%","6.03%"),
    ("腾讯控股","0700.HK","中国","+3.13%","-8.85%","-15.78%","5.99%"),
    ("SK hynix","000660.KS","韩国","+17.24%","+9.49%","+58.05%","5.80%"),
    ("SKSQUARE","402340.KS","韩国","+17.60%","+2.53%","+54.35%","5.47%"),
    ("SamsungElec","005930.KS","韩国","+10.63%","+9.63%","+71.81%","5.13%"),
    ("快手","1024.HK","中国","+0.53%","-26.48%","-29.19%","4.96%"),
    ("SamsungElecMech","009150.KS","韩国","+23.90%","+40.55%","+121.57%","4.01%"),
    ("NAVER","035420.KS","韩国","+2.28%","-8.39%","-15.83%","3.89%"),
    ("联想集团","0992.HK","中国","+4.33%","+6.87%","+9.18%","3.62%"),
    ("华虹公司","1347.HK","中国","+15.94%","+1.49%","+24.16%","3.03%"),
    ("地平线","9660.HK","中国","+8.30%","-5.65%","-16.61%","3.03%"),
    ("商汤集团","0020.HK","中国","+8.65%","-14.47%","-8.64%","2.71%"),
    ("京东健康","6618.HK","中国","+4.78%","-0.19%","-7.65%","2.46%"),
    ("哔哩哔哩-W","9626.HK","中国","+3.14%","-10.46%","-4.11%","2.45%"),
    ("HYOSUNG HEAVY","298040.KS","韩国","+14.84%","+15.29%","+65.08%","1.98%"),
    ("舜宇光学科技","2382.HK","中国","+4.84%","+11.94%","-4.20%","1.94%"),
    ("Kakao","035720.KS","韩国","+5.53%","-7.92%","-20.52%","1.89%"),
]
def cc(v):
    if v.startswith('+'): return '#10b981'
    if v.startswith('-'): return '#ef4444'
    return '#9ca3af'
def wc(v):
    try: return '#ef4444' if float(v.rstrip('%')) >= 5 else '#10b981'
    except: return '#10b981'
h3431_rows = ''.join(
    f"<tr><td style='padding:5px 8px;font-size:12px'>{n}</td><td style='padding:5px 8px;font-size:11px;color:#3b82f6;font-family:monospace'>{c}</td><td style='padding:5px 8px;font-size:11px'>{r}</td><td style='padding:5px 8px;font-size:11px;text-align:right;color:{cc(w1)}'>{w1}</td><td style='padding:5px 8px;font-size:11px;text-align:right;color:{cc(m1)}'>{m1}</td><td style='padding:5px 8px;font-size:11px;text-align:right;color:{cc(ytd)}'>{ytd}</td><td style='padding:5px 8px;font-size:11px;text-align:right;font-weight:700;color:{wc(wgt)}'>{wgt}</td></tr>"
    for n,c,r,w1,m1,ytd,wgt in H3431
)
holdings_cache['3431.HK'] = h3431_rows
print(f"持仓已加载: {len(holdings_cache)} 个")

# ── 颜色映射 ─────────────────────────────────────────────────────────────────
CAT_COLORS = {
    'QDII ETF': ('#3b82f6','#3b82f618'),
    'QDII 联接基金': ('#8b5cf6','#8b5cf618'),
    '港股通ETF': ('#14b8a6','#14b8a618'),
    '港股通ETF-60/40产品': ('#ec4899','#fce7f3'),
    '跨境理财通': ('#f59e0b','#f59e0b18'),
    'A股ETF': ('#3b82f6','#3b82f618'),
    'REITs ETF': ('#f59e0b','#f59e0b18'),
    '个股杠反ETF': ('#ef4444','#fef2f2'),
    '债券-中国': ('#10b981','#d1fae5'),
    '债券-美国': ('#3b82f6','#dbeafe'),
    '加密货币期货ETF': ('#f97316','#ffedd5'),
    '加密货币期货杠反ETF': ('#ef4444','#fef2f2'),
    '商品期货杠反ETF': ('#f59e0b','#fef3c7'),
    '指数杠反ETF': ('#ec4899','#fce7f3'),
    '新兴市场': ('#8b5cf6','#ede9fe'),
    '日本股票ETF': ('#ef4444','#fef2f2'),
    '港股通60/40产品': ('#ec4899','#fce7f3'),
    '美国股票ETF': ('#3b82f6','#dbeafe'),
    '货币市场ETF': ('#10b981','#d1fae5'),
    '跨市场': ('#14b8a6','#d1fae5'),
    '香港': ('#14b8a6','#d1fae5'),
    '香港-美元份额': ('#14b8a6','#d1fae5'),
    'A股': ('#3b82f6','#dbeafe'),
    'REITs': ('#f59e0b','#fef3c7'),
    '商品ETF': ('#f59e0b','#fef3c7'),
}

def chip_color(cat2):
    for key, (c, bg) in CAT_COLORS.items():
        if key in cat2: return c, bg
    return '#6b7280', '#f3f4f6'

# ── 休市日历 ─────────────────────────────────────────────────────────────────
HOLIDAY_ITEMS = (
    "<div style='padding:6px 8px;background:#fef2f2;border-radius:4px;border-left:3px solid #ef4444'><span style='color:#991b1b;font-weight:600'>元旦</span><br><span style='color:#6b7280'>2026-01-01</span></div>"
    "<div style='padding:6px 8px;background:#fef2f2;border-radius:4px;border-left:3px solid #ef4444'><span style='color:#991b1b;font-weight:600'>春节</span><br><span style='color:#6b7280'>2026-01-28~02-04</span></div>"
    "<div style='padding:6px 8px;background:#fff7ed;border-radius:4px;border-left:3px solid #f97316'><span style='color:#9a3412;font-weight:600'>清明</span><br><span style='color:#6b7280'>2026-04-04~06</span></div>"
    "<div style='padding:6px 8px;background:#fff7ed;border-radius:4px;border-left:3px solid #f97316'><span style='color:#9a3412;font-weight:600'>劳动节</span><br><span style='color:#6b7280'>2026-05-01~03</span></div>"
    "<div style='padding:6px 8px;background:#fef9c3;border-radius:4px;border-left:3px solid #eab308'><span style='color:#854d0e;font-weight:600'>端午</span><br><span style='color:#6b7280'>2026-05-31~06-02</span></div>"
    "<div style='padding:6px 8px;background:#fef9c3;border-radius:4px;border-left:3px solid #eab308'><span style='color:#854d0e;font-weight:600'>中秋</span><br><span style='color:#6b7280'>2026-10-01~08</span></div>"
)
HOLIDAY_GRID = "<div style='display:grid;grid-template-columns:1fr 1fr;gap:6px'>" + HOLIDAY_ITEMS + "</div>"

# ── Popup生成 ────────────────────────────────────────────────────────────────
PLACEHOLDER = "<div style='font-size:12px;color:#6b7280;text-align:center;padding:16px 0;background:#f9fafb;border-radius:8px'>持仓数据由富途牛牛等平台提供，详见各平台披露</div>"

def make_holdings_section(rows, date_str):
    if not rows:
        return PLACEHOLDER
    hdr = (
        "<table style='width:100%;border-collapse:collapse'>"
        "<thead><tr style='background:#f0f2f8'>"
        "<th style='padding:5px 8px;text-align:left;font-size:10px;color:#9ca3af'>持仓</th>"
        "<th style='padding:5px 8px;text-align:left;font-size:10px;color:#9ca3af'>代码</th>"
        "<th style='padding:5px 8px;text-align:left;font-size:10px;color:#9ca3af'>地区</th>"
        "<th style='padding:5px 8px;text-align:right;font-size:10px;color:#9ca3af'>1周</th>"
        "<th style='padding:5px 8px;text-align:right;font-size:10px;color:#9ca3af'>1月</th>"
        "<th style='padding:5px 8px;text-align:right;font-size:10px;color:#9ca3af'>YTD</th>"
        "<th style='padding:5px 8px;text-align:right;font-size:10px;color:#9ca3af'>权重</th>"
        "</tr></thead><tbody>" + rows + "</tbody></table>"
    )
    return f"<div style='font-size:11px;color:#9ca3af;margin-bottom:6px'>数据截至: {date_str}</div>" + hdr

def build_popup(code, prod):
    cat2 = prod['cat2']
    color, bg = chip_color(cat2)
    tags = f"<span style='background:{bg};color:{color};border:1px solid {color}33;border-radius:4px;padding:3px 8px;font-size:11px;white-space:nowrap'>{cat2}</span>"
    extra = ''
    if '60/40' in cat2 or code in ['3441.HK','3442.HK','3431.HK']:
        extra = "<div style='margin-top:8px;padding:6px 10px;background:#fce7f3;border-radius:6px;font-size:12px;color:#be185d;font-weight:600'>🔶 港股通60/40产品</div>"

    hrows = holdings_cache.get(code)
    if code == '3431.HK':
        hsection = make_holdings_section(hrows, '2026-04-10')
    elif hrows:
        hsection = make_holdings_section(hrows, '2026-04-17')
    else:
        hsection = PLACEHOLDER

    holdings_block = (
        "<div style='margin-top:16px;padding-top:16px;border-top:1px solid #e5e7eb'>"
        "<div style='font-size:13px;font-weight:600;color:#374151;margin-bottom:10px'>📊 持仓明细</div>"
        "<div style='margin-top:12px'>" + hsection + "</div>"
        "</div>"
    )
    holiday_block = (
        "<div style='margin-top:16px;padding-top:16px;border-top:1px solid #e5e7eb'>"
        "<div style='font-size:13px;font-weight:600;color:#374151;margin-bottom:10px'>📅 休市日历（2025-2026）</div>"
        "<div style='font-size:11px;color:#374151'>" + HOLIDAY_GRID + "</div>"
        "</div>"
    )
    mainland_tag = '🇹🇨 内地直接可买' if prod['cat1'] == '内地可买' else '🌍 境外账户可买'
    return (
        f'<div>'
        f'<div style="display:flex;align-items:flex-start;justify-content:space-between;gap:12px;margin-bottom:10px">'
        f'<div><div style="font-size:16px;font-weight:700">{prod["name"]}</div>'
        f'<div style="font-size:13px;color:#3b82f6;margin-top:2px;font-family:monospace">{code}</div></div>'
        f'{tags}</div>'
        f'<div style="font-size:12px;color:#6b7280;margin-bottom:8px">{mainland_tag}</div>'
        + extra + holdings_block + holiday_block +
        '</div>'
    )

popup_data = {code: build_popup(code, prod) for code, prod in products.items()}

# ── 芯片HTML ─────────────────────────────────────────────────────────────────
def chip(code, name):
    return f"<span class='prod-chip' data-code='{code}' onclick=\"showPopup('{code}')\">{name}</span>"

def section(cat2, prods_list, color, bg):
    chips = ''.join(chip(code, p['name']) for code, p in prods_list)
    return (
        f"<div style='margin-bottom:20px'>"
        f"<div style='display:flex;align-items:center;gap:8px;margin-bottom:8px'>"
        f"<span style='background:{bg};color:{color};border:1px solid {color}33;border-radius:4px;padding:2px 8px;font-size:11px;white-space:nowrap'>{cat2}</span>"
        f"<span style='font-size:11px;color:#9ca3af'>{len(prods_list)}只</span></div>"
        f"<div style='display:flex;flex-wrap:wrap;gap:6px'>{chips}</div></div>"
    )

def build_mainland():
    groups = {}
    for code, p in mainland_prods.items():
        groups.setdefault(p['cat2'], []).append((code, p))
    ORDER = ['QDII ETF','QDII 联接基金','港股通ETF','港股通ETF-60/40产品']
    parts = []
    done = set()
    for cat2 in ORDER:
        if cat2 in groups:
            color, bg = chip_color(cat2)
            parts.append(section(cat2, groups[cat2], color, bg))
            done.add(cat2)
    # 大湾区理财通 merge
    gp = [g for g in groups if '跨境理财通' in g and g not in done]
    if gp:
        all_gp = []
        for g in gp:
            all_gp.extend(groups[g])
            done.add(g)
        if all_gp:
            color, bg = chip_color('跨境理财通')
            parts.append(section('🌐 大湾区理财通', all_gp, color, bg))
    for cat2, sub in groups.items():
        if cat2 not in done:
            color, bg = chip_color(cat2)
            parts.append(section(cat2, sub, color, bg))
    return ''.join(parts)

def build_overseas():
    groups = {}
    for code, p in overseas_prods.items():
        groups.setdefault(p['cat2'], []).append((code, p))
    ORDER = [
        '港股指数','港股主题','港股通60/40产品','区域性',
        'A股','美国股票ETF','日本股票ETF',
        'REITs ETF','REITs','债券-中国','债券-美国',
        '货币市场ETF','货币市场ETF-美元份额',
        '跨市场','香港','香港-美元份额',
        '新兴市场',
        '加密货币期货ETF','加密货币期货杠反ETF',
        '商品期货杠反ETF','商品ETF',
        '指数杠反ETF','个股杠反ETF','个股杠反ETF-美元份额',
    ]
    parts = []
    done = set()
    for cat2 in ORDER:
        if cat2 in groups and cat2 not in done:
            color, bg = chip_color(cat2)
            parts.append(section(cat2, groups[cat2], color, bg))
            done.add(cat2)
    for cat2, sub in groups.items():
        if cat2 not in done:
            color, bg = chip_color(cat2)
            parts.append(section(cat2, sub, color, bg))
    return ''.join(parts)

# ── 生成HTML ─────────────────────────────────────────────────────────────────
CSS = """:root{--bg:#0d1117;--card:#161b22;--border:#30363d;--text:#e6edf3;--dim:#8b949e;--green:#3fb950;--red:#f85149;--accent:#ffa657;--blue:#58a6ff}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'PingFang SC','Microsoft YaHei',-apple-system,sans-serif;background:var(--bg);color:var(--text);min-height:100vh}
.header{background:var(--card);border-bottom:1px solid var(--border);padding:18px 36px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:12px}
.header h1{font-size:22px;font-weight:700}
.header-right{font-size:12px;color:var(--dim)}
.nav{display:flex;gap:20px;padding:12px 36px;background:var(--card);border-bottom:1px solid var(--border)}
.nav a{color:var(--dim);text-decoration:none;font-size:14px;padding:6px 12px;border-radius:6px;transition:.2s}
.nav a:hover,.nav a.active{color:var(--text);background:#1c2128}
.channel-bar{padding:16px 36px;background:var(--card);border-bottom:1px solid var(--border)}
.channel-label{font-size:12px;color:var(--dim);margin-bottom:10px;text-transform:uppercase;letter-spacing:.5px}
.channel-tabs{display:flex;gap:10px;flex-wrap:wrap}
.chan-btn{background:#21262d;border:1px solid #30363d;color:var(--text);padding:10px 20px;border-radius:8px;cursor:pointer;font-size:14px;transition:.2s}
.chan-btn:hover{border-color:var(--dim)}
.chan-btn.active{background:#238636;border-color:#238636;color:#fff}
.tab-content{display:none;padding:24px 36px;max-width:1400px;margin:0 auto}
.tab-content.active{display:block}
.prod-chip{display:inline-block;padding:6px 12px;background:#1c2128;border:1px solid #30363d;border-radius:20px;font-size:13px;cursor:pointer;transition:.2s;margin-bottom:4px}
.prod-chip:hover{background:#2d333b;border-color:var(--dim)}
.ch-intro{font-size:13px;color:var(--dim);padding:16px 36px;line-height:1.6}
.ch-table{padding:0 36px 24px}
.ch-table table{width:100%;border-collapse:collapse;font-size:13px}
.ch-table th{text-align:left;padding:10px 14px;background:#1c2128;border-bottom:1px solid var(--border);color:var(--dim);font-size:11px;text-transform:uppercase;letter-spacing:.5px;white-space:nowrap}
.ch-table td{padding:10px 14px;border-bottom:1px solid #21262d;font-size:13px;white-space:nowrap}
.ch-table tr:hover td{background:#161b22}
.ch-table td:first-child{color:var(--text);font-weight:600}
/* Popup */
#popupOverlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,0.6);z-index:1000;align-items:center;justify-content:center;padding:20px}
#popupOverlay.show{display:flex}
.popup-box{background:#fff;border-radius:16px;max-width:680px;width:100%;max-height:85vh;overflow-y:auto;box-shadow:0 20px 60px rgba(0,0,0,0.3)}
.popup-header{padding:16px 20px;border-bottom:1px solid #f0f2f8;display:flex;align-items:center;justify-content:space-between;flex-shrink:0}
.popup-header h3{font-size:16px;font-weight:700;color:#1a1d2e}
.popup-close{width:32px;height:32px;border-radius:50%;border:none;background:#f0f2f8;cursor:pointer;font-size:18px;display:flex;align-items:center;justify-content:center;color:#6b7280;transition:all 0.15s}
.popup-close:hover{background:#e2e5ef;color:#1a1d2e}
.popup-body{padding:20px}
@media(max-width:600px){.chan-btn{font-size:12px;padding:10px 12px}}"""

CHANNEL_ROWS = (
    "<tr><td style='color:#1a1d2e;font-weight:600'>QDII ETF</td><td style='color:#6b7280'>香港/美国/日本等</td><td style='color:#6b7280'>100元起</td><td style='color:#6b7280'>有额度限制</td><td style='color:#6b7280'>香港/美国/日本等市场ETF</td></tr>"
    "<tr><td style='color:#1a1d2e;font-weight:600'>QDII 联接基金</td><td style='color:#6b7280'>香港/美国/日本等</td><td style='color:#6b7280'>10元起</td><td style='color:#6b7280'>有额度限制</td><td style='color:#6b7280'>挂钩海外ETF的基金</td></tr>"
    "<tr><td style='color:#1a1d2e;font-weight:600'>港股通ETF</td><td style='color:#6b7280'>香港</td><td style='color:#6b7280'>几乎无门槛</td><td style='color:#6b7280'>每日520亿元</td><td style='color:#6b7280'>香港股票ETF（南向）</td></tr>"
    "<tr><td style='color:#1a1d2e;font-weight:600'>大湾区理财通</td><td style='color:#6b7280'>香港/澳门</td><td style='color:#6b7280'>10元起</td><td style='color:#6b7280'>个人户每年30万</td><td style='color:#6b7280'>基金/债券/存款/理财</td></tr>"
    "<tr><td style='color:#1a1d2e;font-weight:600'>境外证券账户</td><td style='color:#6b7280'>全球</td><td style='color:#6b7280'>需开境外证券账户</td><td style='color:#6b7280'>无额度限制</td><td style='color:#6b7280'>全球各类型资产和市场ETF</td></tr>"
)

mainland_html = build_mainland()
overseas_html = build_overseas()
popup_json = json.dumps(popup_data, ensure_ascii=False)

parts = []
parts.append("<!DOCTYPE html><html lang='zh-CN'><head><meta charset='UTF-8'><meta name='viewport' content='width=device-width, initial-scale=1.0'><title>全球ETF投资地图</title><style>")
parts.append(CSS)
parts.append("</style></head><body>")
parts.append("<header class='header'><h1>🌏 全球ETF投资地图</h1><div class='header-right'>数据更新: %s</div></header>" % TODAY)
parts.append("<nav class='nav'><a href='index.html' class='active'>🗺️ 全球ETF地图</a><a href='products.html'>📋 产品总览</a></nav>")
parts.append("<div class='channel-bar'><div class='channel-label'>选择投资渠道</div><div class='channel-tabs'>")
parts.append("<button class='chan-btn active' id='chan-mainland' onclick=\"switchChannel('mainland')\">🇹🇨 内地投资者直接可买 <span style='opacity:0.65;font-size:11px;font-weight:400'>(%d)</span></button>" % len(mainland_prods))
parts.append("<button class='chan-btn' id='chan-overseas' onclick=\"switchChannel('overseas')\">🌍 境外账户可买 <span style='opacity:0.65;font-size:11px;font-weight:400'>(%d)</span></button>" % len(overseas_prods))
parts.append("</div></div>")
parts.append("<div id='tab-mainland' class='tab-content active'>%s</div>" % mainland_html)
parts.append("<div id='tab-overseas' class='tab-content'>%s</div>" % overseas_html)
parts.append("<div class='ch-intro'><strong>🌐 四大跨境投资渠道对比</strong> — 不同渠道在投资门槛、额度限制、可投标的范围等方面各有不同，适合不同需求的投资者。</div>")
parts.append("<div class='ch-table'><table><thead><tr><th style='width:150px'>渠道</th><th style='width:80px'>可参与地区</th><th style='width:180px'>投资门槛</th><th style='width:150px'>额度限制</th><th>资金投向</th></tr></thead><tbody>")
parts.append(CHANNEL_ROWS)
parts.append("</tbody></table></div>")
parts.append("<div id='popupOverlay' onclick='closeOnBg(event)'><div class='popup-box'><div class='popup-header'><h3 id='popupTitle'></h3><button class='popup-close' onclick='closePopup()'>×</button></div><div class='popup-body' id='popupBody'></div></div></div>")
parts.append("<script>var POPUP_DATA=" + popup_json + ";")
parts.append("function switchChannel(ch){document.getElementById('chan-mainland').classList.toggle('active',ch==='mainland');document.getElementById('chan-overseas').classList.toggle('active',ch==='overseas');document.getElementById('tab-mainland').classList.toggle('active',ch==='mainland');document.getElementById('tab-overseas').classList.toggle('active',ch==='overseas');}")
parts.append("function showPopup(code){var d=POPUP_DATA[code];if(!d)return;var ov=document.getElementById('popupOverlay');document.getElementById('popupBody').innerHTML=d;ov.classList.add('show');}")
parts.append("function closePopup(){document.getElementById('popupOverlay').classList.remove('show');}")
parts.append("function closeOnBg(e){if(e.target===document.getElementById('popupOverlay'))closePopup();}")
parts.append("document.addEventListener('keydown',function(e){if(e.key==='Escape')closePopup();});")
parts.append("</script></body></html>")

html = ''.join(parts)

with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"Done! {len(html)} bytes -> {OUTPUT_FILE}")
print(f"Products: {len(products)} | Mainland: {len(mainland_prods)} | Overseas: {len(overseas_prods)}")
