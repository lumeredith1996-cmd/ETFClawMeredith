
#!/usr/bin/env python3
import openpyxl, json, re
from collections import defaultdict

wb = openpyxl.load_workbook('/root/.openclaw/media/inbound/all产品_20260419.xlsx')
ws = wb['products_all']
products = []
for row in ws.iter_rows(values_only=True):
    if row[0] is None or row[0] == '一级分类': continue
    products.append({'cat1':str(row[0]).strip(),'cat2':str(row[1]).strip(),'code':str(row[2]).strip() if row[2] else '','name':str(row[3]).strip() if row[3] else '','note':str(row[4]).strip() if row[4] else ''})

CROSSBORDER_CATS={'跨境理财通（大湾区9市）-权益','跨境理财通（大湾区9市）-货币','跨境理财通（大湾区9市）-固定收益','跨境理财通（大湾区9市）-基金互认'}
for p in products:
    if p['cat1']=='内地可买' and p['cat2'] not in CROSSBORDER_CATS: p['channel']='mainland'
    elif p['cat1']=='内地可买' and p['cat2'] in CROSSBORDER_CATS: p['channel']='crossborder'
    else: p['channel']='overseas'

CAT2_DISPLAY={'QDII ETF':'QDII ETF','QDII 联接基金':'QDII 联接基金','港股通ETF':'港股通ETF','港股通ETF-60/40产品':'港股通60/40产品','跨境理财通（大湾区9市）-权益':'跨境理财通·权益','跨境理财通（大湾区9市）-货币':'跨境理财通·货币','跨境理财通（大湾区9市）-固定收益':'跨境理财通·固收','跨境理财通（大湾区9市）-基金互认':'基金互认','个股杠反ETF':'个股杠反ETF','指数杠反ETF':'指数杠反ETF','加密货币期货ETF':'加密期货ETF','加密货币期货杠反ETF':'加密期货杠反','商品期货杠反ETF':'商品期货杠反','债券ETF-中国':'债券-中国','债券ETF-美国':'债券-美国','货币市场ETF':'货币市场ETF','REITs ETF':'REITs ETF','香港':'香港股票ETF','香港-美元份额':'香港-美元','美国':'美国股票ETF','A股':'A股ETF','新兴市场':'新兴市场','日本':'日本股票ETF','跨市场':'跨市场'}
CAT2_COLOR={'QDII ETF':'#3b82f6','QDII 联接基金':'#6366f1','港股通ETF':'#10b981','港股通ETF-60/40产品':'#14b8a6','跨境理财通（大湾区9市）-权益':'#f59e0b','跨境理财通（大湾区9市）-货币':'#fbbf24','跨境理财通（大湾区9市）-固定收益':'#f97316','跨境理财通（大湾区9市）-基金互认':'#eab308','个股杠反ETF':'#ec4899','指数杠反ETF':'#ef4444','加密货币期货ETF':'#fbbf24','加密货币期货杠反ETF':'#f59e0b','商品期货杠反ETF':'#eab308','债券ETF-中国':'#8b5cf6','债券ETF-美国':'#6366f1','货币市场ETF':'#10b981','REITs ETF':'#14b8a6','香港':'#ef4444','香港-美元份额':'#dc2626','美国':'#3b82f6','A股':'#8b5cf6','新兴市场':'#f59e0b','日本':'#ec4899','跨市场':'#14b8a6'}

def parse_holdings(path):
    try:
        wb2=openpyxl.load_workbook(path);ws2=wb2['Performance Report'];rows=list(ws2.iter_rows(values_only=True))
        data_date=rows[1][0].replace('数据截至日期: ','') if len(rows)>1 and rows[1][0] else ''
        holdings=[];idx_section=False
        for r in rows[3:]:
            if r[0] and '主要市场' in str(r[0]): idx_section=True;continue
            if r[0] in('前二十大大持仓','前二十大持仓','前十大持仓') or(r[0] and '持仓' in str(r[0])): continue
            if idx_section and r[1] is None: continue
            if r[1] and r[1]!='代码':
                name=str(r[0]) if r[0] else '';ticker=str(r[1]) if r[1] else '';region=str(r[2]) if r[2] else ''
                try:
                    wgt=float(r[9]) if r[9] is not None else 0;w1=float(r[4]) if r[4] is not None else 0
                    m1=float(r[5]) if r[5] is not None else 0;ytd=float(r[6]) if r[6] is not None else 0
                except: continue
                if ticker and name and wgt>0: holdings.append({'name':name,'ticker':ticker,'region':region,'w1':w1,'m1':m1,'ytd':ytd,'weight':wgt})
        return data_date,holdings
    except: return '',[]

holdings_db={}
for code,path in {'513730.SH':'/root/.openclaw/media/inbound/\xe4\xb8\x9c\xe5\x9c\xb0\xe7\xa7\x91\xe6\x8a\x80ETF_2026-04-10---bb2540ac-8510-4621-a4f2-fecd4cf4ff72.xlsx','159687.SZ':'/root/.openclaw/media/inbound/\xe4\xba\x9a\xe5\xa4\xaa\xe7\xb2\xbe\xe9\x80\x89ETF_2026-04-10---73e27b36-185e-4f17-8d69-39c201057957.xlsx','3441.HK':'/root/.openclaw/media/inbound/\xe4\xb8\x9c\xe8\xa5\xbf\xe7\xb2\xbe\xe9\x80\x89_2026-04-10---43206195-743b-4dd8-a066-d2bbaecc5d58.xlsx','3442.HK':'/root/.openclaw/media/inbound/\xe6\xb8\xaf\xe7\xbe\x8e\xe7\xa7\x91\xe6\x8a\x80_2026-04-10---02a799fe-55e9-45af-b620-288cd5b1b0aa.xlsx','2830.HK':'/root/.openclaw/media/inbound/\xe6\xb2\x99\xe7\x89\xb9_2026-04-13---baa6fc66-98d6-428f-b8f7-6db42705cc21.xlsx'}.items():
    date,h=parse_holdings(path)
    if h: holdings_db[code]={'date':date,'holdings':h}

def expand_date_range(ds,r):
    if '－' not in ds and '-' not in ds and '至' not in ds: return [(ds,r)]
    sep='－' if '－' in ds else('-' if '-' in ds else '至')
    p=ds.split(sep);s,e=p[0].strip(),p[1].strip() if len(p)>1 else s
    ms=int(re.search(r'(\d+)月',s).group(1));ds2=int(re.search(r'(\d+)日',s).group(1))
    me=int(re.search(r'(\d+)月',e).group(1));de=int(re.search(r'(\d+)日',e).group(1))
    if ms==me: return [(f'{ms}月{d}日',r) for d in range(ds2,de+1)]
    return [(s,r),(e,r)]

HOLIDAYS={}
for code,info in {'513730.SH':{'raw':[('1月16日',''),('1月19日',''),('3月3日',''),('3月18日-3月20日',''),('3月23日-3月24日',''),('4月2日-4月3日',''),('4月13日-4月15日',''),('5月14日-5月15日',''),('5月25日-5月29日',''),('6月1日',''),('6月3日',''),('6月16日',''),('7月3日',''),('7月28日-7月29日',''),('8月7日',''),('8月10日',''),('8月12日',''),('8月17日',''),('8月25日',''),('9月7日',''),('10月13日',''),('10月23日',''),('11月6日',''),('11月9日',''),('11月26日',''),('12月7日',''),('12月10日',''),('12月24日',''),('12月25日',''),('12月31日','')],'note':'泰国2026年休市日历'},'159687.SZ':{'raw':[('1月12日','日本成人节'),('1月19日','美国马丁·路德·金节'),('2月11日','日本建国纪念日'),('3月20日','日本春分节'),('4月3日','香港/美国/新加坡耶稣受难日'),('4月7日','香港复活节假期'),('4月29日','日本昭和日'),('5月6日','日本宪法纪念日假期'),('5月25日','香港佛诞节/美国阵亡将士纪念日'),('5月27日','新加坡哈芝节'),('6月1日','新加坡卫塞节假期'),('7月1日','香港特区成立纪念日'),('7月3日','美国独立日假期'),('7月20日','日本海之日'),('8月10日','新加坡国庆日假期'),('8月11日','日本山之日'),('9月7日','美国劳动节'),('9月21日','日本敬老日'),('9月22日',''),('9月23日','日本秋分节假期'),('10月12日','日本体育日'),('10月19日','香港重阳节假期'),('11月3日','日本文化日'),('11月9日','新加坡屠妖节假期'),('11月23日','日本劳动感谢日'),('11月26日','美国感恩节'),('12月24日','港新圣诞节前日'),('12月25日','港美新圣诞节'),('12月31日','港新新年假期前日/日本银行假日')],'note':'亚太精选ETF 2026年休市日历（日本/美国/香港/新加坡）'}}.items():
    HOLIDAYS[code]={'holidays':[],'note':info['note']}
    for item in info['raw']: HOLIDAYS[code]['holidays'].extend(expand_date_range(item[0],item[1] if len(item)>1 else ''))

def fp(v): return f'{v:+.2f}%' if v is not None else '—'
def esc(s): return s.replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;').replace("'",'&#39;')
LF_UNDERLYING={'021189.OF':'159687.SZ','020515.OF':'513730.SH','020516.OF':'513730.SH'}

def holiday_section(code):
    info=HOLIDAYS.get(code)
    if not info: return ''
    by_month=defaultdict(list);MO=['1月','2月','3月','4月','5月','6月','7月','8月','9月','10月','11月','12月']
    for ds,r in info['holidays']:
        m=ds[:ds.index('月')+1] if '月' in ds else ds
        by_month[m].append((ds,r))
    cal=''
    for m in MO:
        if m not in by_month: continue
        days=''.join('<div style="padding:3px 6px;font-size:12px;border-radius:4px;background:#fff;border:1px solid #e2e5ef;min-width:80px;margin:2px"><div style="font-weight:600;color:#1a1d2e">'+esc(d)+'</div><div style="font-size:10px;color:#6b7280">'+esc(r)+'</div></div>' for d,r in by_month[m])
        cal+='<div style="margin-bottom:8px"><div style="font-size:11px;color:#9ca3af;margin-bottom:4px;font-weight:600">'+m+'</div><div style="display:flex;flex-wrap:wrap;gap:4px">'+days+'</div></div>'
    return '<div style="margin-top:14px;padding:12px;background:#f8f9ff;border-radius:10px;border:1px solid #e2e5ef"><div style="font-size:12px;font-weight:700;color:#1a1d2e;margin-bottom:8px">📅 2026年休市日历</div><div style="font-size:11px;color:#6b7280;margin-bottom:10px">'+esc(info['note'])+'</div>'+cal+'</div>'

def holdings_section(h):
    if not h: return ''
    rows=''.join('<tr><td style="padding:5px 8px;font-size:12px">'+esc(s['name'])+'</td><td style="padding:5px 8px;font-size:11px;color:#3b82f6;font-family:monospace">'+esc(s['ticker'])+'</td><td style="padding:5px 8px;font-size:11px">'+esc(s['region'])+'</td><td style="padding:5px 8px;font-size:11px;text-align:right">'+fp(s['w1'])+'</td><td style="padding:5px 8px;font-size:11px;text-align:right">'+fp(s['m1'])+'</td><td style="padding:5px 8px;font-size:11px;text-align:right">'+fp(s['ytd'])+'</td><td style="padding:5px 8px;font-size:11px;text-align:right;font-weight:700;color:#10b981">'+f"{s['weight']:.2f}%"+'</td></tr>' for s in h['holdings'][:10])
    return '<div style="margin-top:12px"><div style="font-size:11px;color:#9ca3af;margin-bottom:6px">数据截至: '+esc(h['date'])+'</div><table style="width:100%;border-collapse:collapse"><thead><tr style="background:#f0f2f8"><th style="padding:5px 8px;text-align:left;font-size:10px;color:#9ca3af">持仓</th><th style="padding:5px 8px;text-align:left;font-size:10px;color:#9ca3af">代码</th><th style="padding:5px 8px;text-align:left;font-size:10px;color:#9ca3af">地区</th><th style="padding:5px 8px;text-align:right;font-size:10px;color:#9ca3af">1周</th><th style="padding:5px 8px;text-align:right;font-size:10px;color:#9ca3af">1月</th><th style="padding:5px 8px;text-align:right;font-size:10px;color:#9ca3af">YTD</th><th style="padding:5px 8px;text-align:right;font-size:10px;color:#9ca3af">权重</th></tr></thead><tbody>'+rows+'</tbody></table></div>'

def popup_html(code,name,cat2,note,channel):
    c2d=CAT2_DISPLAY.get(cat2,cat2);c2c=CAT2_COLOR.get(cat2,'#6b7280')
    chan_lbl={'mainland':'🇨🇳 内地直接可买','overseas':'🌍 境外账户可买','crossborder':'🌐 跨境理财通'}
    underlying=LF_UNDERLYING.get(code);dc=underlying if underlying else code
    h=holdings_db.get(dc);hh=holiday_section(dc)
    nh='' if not note or note=='None' else '<div style="margin-top:8px;padding:6px 10px;background:#fffbeb;border-radius:6px;font-size:12px;color:#d97706">⚠️ '+esc(note)+'</div>'
    lfe=''
    if underlying: lfe='<div style="margin-top:8px;padding:8px 12px;background:#eff6ff;border-radius:6px;font-size:12px;color:#1d4ed8"><strong>底层标的:</strong> '+underlying+'（持仓与休市日历同上）</div>'
    if note and '限额' in note: lfe+='<div style="margin-top:6px;padding:6px 10px;background:#fef3c7;border-radius:6px;font-size:12px;color:#92400e"><strong>⚠️ 限额:</strong> '+esc(note)+'</div>'
    hs=holdings_section(h)
    return '<div><div style="display:flex;align-items:flex-start;justify-content:space-between;gap:12px;margin-bottom:10px"><div><div style="font-size:16px;font-weight:700">'+esc(name)+'</div><div style="font-size:13px;color:#3b82f6;margin-top:2px;font-family:monospace">'+esc(code)+'</div></div><span style="background:'+c2c+'18;color:'+c2c+';border:1px solid '+c2c+'33;border-radius:4px;padding:3px 8px;font-size:11px;white-space:nowrap">'+esc(c2d)+'</span></div><div style="font-size:12px;color:#6b7280;margin-bottom:6px">'+chan_lbl.get(channel,channel)+'</div>'+nh+lfe+hs+hh+'</div>'

CHANNEL_INFO=[('🌐','QDII','全国','基本无门槛','每家基金管理人额度有限','境外股票和债券市场'),('🔗','港股通ETF（互联互通）','全国','50万资产开通港股通账户','每日额度420亿港元','港股（含ETF，不超过40%非港股仓位）'),('🌏','大湾区跨境理财通','大湾区9市','大湾区居民，有资产门槛','总额度1500亿人民币','境外存款、债券、投资于大中华市场的基金'),('🏦','境外账户直投','全球','需开境外证券账户','无额度限制','全球各类型资产和市场')]
chan_rows=''.join('<tr><td style="padding:12px 16px;border-bottom:1px solid #f0f2f8;vertical-align:top"><span style="font-size:18px;margin-right:8px">'+i+'</span><strong style="font-size:13px">'+esc(l)+'</strong></td><td style="padding:12px 16px;border-bottom:1px solid #f0f2f8;font-size:12px;color:#6b7280;vertical-align:top">'+esc(r)+'</td><td style="padding:12px 16px;border-bottom:1px solid #f0f2f8;font-size:12px;color:#6b7280;vertical-align:top">'+esc(t)+'</td><td style="padding:12px 16px;border-bottom:1px solid #f0f2f8;font-size:12px;color:#6b7280;vertical-align:top">'+esc(q)+'</td><td style="padding:12px 16px;border-bottom:1px solid #f0f2f8;font-size:12px;vertical-align:top">'+esc(x)+'</td></tr>' for i,l,r,t,q,x in CHANNEL_INFO)

OVERS={'香港','A股','美国','日本','新兴市场','跨市场','香港-美元份额','指数杠反ETF','个股杠反ETF','加密货币期货ETF','加密货币期货杠反ETF','商品期货杠反ETF','债券ETF-中国','债券ETF-美国','货币市场ETF','货币市场ETF-美元份额','REITs ETF'}
mainland=[p for p in products if p['channel']=='mainland' and p['cat2'] not in OVERS]
overseas=[p for p in products if p['channel']=='overseas']
cross=[p for p in products if p['channel']=='crossborder']

def grp(ps):
    g=defaultdict(list)
    for p in ps: g[p['cat2']].append(p)
    return sorted(g.items(),key=lambda x:CAT2_DISPLAY.get(x[0],x[0]))

def mkchip(p):
    c=esc(p['code']);n=esc(p['name'])
    return '<span class="prod-chip" data-code="'+c+'" onclick="showPopup(\''+c+'\')">'+n+'</span>'

def mkcat(cat2,prods,color):
    cat2_d=CAT2_DISPLAY.get(cat2,cat2)
    chips=''.join(mkchip(p) for p in prods)
    return '<div style="margin-bottom:20px"><div style="display:flex;align-items:center;gap:8px;margin-bottom:8px"><span style="background:'+color+'18;color:'+color+';border:1px solid '+color+'33;border-radius:4px;padding:2px 8px;font-size:11px;white-space:nowrap">'+esc(cat2_d)+'</span><span style="font-size:11px;color:#9ca3af">'+str(len(prods))+'只</span></div><div style="display:flex;flex-wrap:wrap;gap:6px">'+chips+'</div></div>'

def mktab(mg,og,cg):
    mh=''.join(mkcat(cat2,prods,CAT2_COLOR.get(cat2,'#6b7280')) for cat2,prods in mg)
    if cg:
        cb=''.join(mkcat(cat2,prods,'#f59e0b') for cat2,prods in cg)
        mh+='<div style="margin-top:28px;padding-top:20px;border-top:2px dashed #e2e5ef"><div style="font-size:13px;font-weight:700;color:#92400e;margin-bottom:14px">🌐 大湾区跨境理财通</div>'+cb+'</div>'
    oh=''.join(mkcat(cat2,prods,CAT2_COLOR.get(cat2,'#6b7280')) for cat2,prods in og)
    return mh,oh

mg,og,cg=grp(mainland),grp(overseas),grp(cross)
mh,oh=mktab(mg,og,cg)
cnt_m=len(set(p['code'] for p in mainland)|set(p['code'] for p in cross))
cnt_o=len(set(p['code'] for p in overseas))
popup={}
for p in products: popup[p['code']]=popup_html(p['code'],p['name'],p['cat2'],p['note'],p['channel'])
pj=json.dumps(popup,ensure_ascii=False)

CSS="""*{box-sizing:border-box;margin:0;padding:0}body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f5f6fa;color:#1a1d2e;min-height:100vh}.header{background:#1a1d2e;color:#fff;padding:18px 24px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px}.header h1{font-size:20px;font-weight:700;display:flex;align-items:center;gap:8px}.header-right{font-size:12px;opacity:0.55}.nav{display:flex;gap:8px;padding:12px 24px;background:#fff;border-bottom:1px solid #e2e5ef;flex-wrap:wrap}.nav a{color:#6b7280;text-decoration:none;font-size:14px;padding:6px 14px;border-radius:6px;transition:all 0.15s;display:inline-block}.nav a:hover{background:#f0f2f8;color:#1a1d2e}.nav a.active{background:#1a1d2e;color:#fff}.channel-bar{padding:20px 24px 0;background:#fff;border-bottom:2px solid #e2e5ef}.channel-label{font-size:11px;color:#9ca3af;font-weight:700;letter-spacing:0.1em;text-transform:uppercase;margin-bottom:10px}.channel-tabs{display:flex;gap:8px;flex-wrap:wrap}.chan-btn{flex:1;min-width:160px;max-width:260px;padding:12px 16px;border:2px solid #e2e5ef;border-bottom:none;border-radius:10px 10px 0 0;font-size:13px;font-weight:700;cursor:pointer;background:#f8f9fc;color:#6b7280;transition:all 0.15s;text-align:center}.chan-btn:hover{background:#f0f2f8;color:#1a1d2e}.chan-btn.active{background:#1a1d2e;color:#fff;border-color:#1a1d2e}.tab-content{display:none;padding:20px 24px}.tab-content.active{display:block}.prod-chip{background:#f0f2f8;border-radius:6px;padding:7px 12px;font-size:13px;cursor:pointer;display:inline-block;transition:all 0.15s;margin:3px;border:1px solid transparent}.prod-chip:hover{background:#e2e5ef;color:#1a1d2e}.ch-intro{padding:16px 24px;background:#fff;border-top:1px solid #e2e5ef;font-size:13px;color:#6b7280;line-height:1.6}.ch-table{width:100%;border-collapse:collapse;background:#fff;border-radius:0 0 10px 10px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.06);border:1px solid #e2e5ef;border-top:none}.ch-table thead th{background:#1a1d2e;color:#fff;font-size:11px;padding:10px 16px;text-align:left}.ch-table td{font-size:13px}.ch-table tr:hover{background:#f8fbff}#popupOverlay{display:none;position:fixed;inset:0;z-index:1000;background:rgba(0,0,0,0.5);backdrop-filter:blur(4px);justify-content:center;align-items:center;padding:20px}#popupOverlay.show{display:flex}.popup-box{background:#fff;border-radius:16px;max-width:680px;width:100%;max-height:85vh;overflow-y:auto;box-shadow:0 20px 60px rgba(0,0,0,0.3)}.popup-header{padding:16px 20px;border-bottom:1px solid #f0f2f8;display:flex;align-items:center;justify-content:space-between;flex-shrink:0}.popup-header h3{font-size:16px;font-weight:700}.popup-close{width:32px;height:32px;border-radius:50%;border:none;background:#f0f2f8;cursor:pointer;font-size:18px;display:flex;align-items:center;justify-content:center;color:#6b7280;transition:all 0.15s;flex-shrink:0}.popup-close:hover{background:#e2e5ef;color:#1a1d2e}.popup-body{padding:20px}@media(max-width:600px){.chan-btn{font-size:12px;padding:10px 12px}}"""

with open('/root/.openclaw/workspace/etf-map-site/index.html','w',encoding='utf-8') as f:
    f.write('<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"><title>全球ETF投资地图 — 投资工具箱</title><style>'+CSS+'</style></head><body>')
    f.write('<header class="header"><h1>🌏 全球ETF投资地图</h1><div class="header-right">数据更新: 2026-04-19</div></header>')
    f.write('<nav class="nav"><a href="index.html" class="active">🗺️ 全球ETF地图</a><a href="products.html">📋 产品总览</a></nav>')
    f.write('<div class="channel-bar"><div class="channel-label">选择投资渠道</div><div class="channel-tabs">')
    f.write("<button class=\"chan-btn active\" id=\"chan-mainland\" onclick=\"switchChannel('mainland')\">🇹🇨 内地投资者直接可买 <span style=\"opacity:0.65;font-size:11px;font-weight:400\">("+str(cnt_m)+")</span></button>")
    f.write("<button class=\"chan-btn\" id=\"chan-overseas\" onclick=\"switchChannel('overseas')\">🌍 境外账户可买 <span style=\"opacity:0.65;font-size:11px;font-weight:400\">("+str(cnt_o)+")</span></button>")
    f.write('</div></div>')
    f.write('<div id="tab-mainland" class="tab-content active">'+mh+'</div>')
    f.write('<div id="tab-overseas" class="tab-content">'+oh+'</div>')
    f.write('<div class="ch-intro"><strong>🌐 四大跨境投资渠道对比</strong> — 不同渠道在投资门槛、额度限制、可投标的范围等方面各有不同，适合不同需求的投资者。</div>')
    f.write('<div style="padding:0 24px 24px"><table class="ch-table"><thead><tr><th style="width:150px">渠道</th><th style="width:80px">可参与地区</th><th style="width:180px">投资门槛</th><th style="width:150px">额度限制</th><th>资金投向</th></tr></thead><tbody>'+chan_rows+'</tbody></table></div>')
    f.write('<div id="popupOverlay" onclick="closeOnBg(event)"><div class="popup-box"><div class="popup-header"><h3>📋 产品详情</h3><button class="popup-close" onclick="closePopup()">×</button></div><div class="popup-body" id="popupBody"></div></div></div>')
    f.write('<script>var POPUP_DATA='+pj+';')
    f.write("function switchChannel(ch){document.querySelectorAll('.chan-btn').forEach(function(b){b.classList.remove('active')});document.getElementById('chan-'+ch).classList.add('active');document.querySelectorAll('.tab-content').forEach(function(t){t.classList.remove('active')});document.getElementById('tab-'+ch).classList.add('active')} ")
    f.write("function showPopup(code){var d=POPUP_DATA[code];if(!d)return;document.getElementById('popupBody').innerHTML=d;document.getElementById('popupOverlay').classList.add('show');document.body.style.overflow='hidden'} ")
    f.write("function closePopup(){document.getElementById('popupOverlay').classList.remove('show');document.body.style.overflow=''} ")
    f.write("function closeOnBg(e){if(e.target===document.getElementById('popupOverlay'))closePopup()} ")
    f.write("document.addEventListener('keydown',function(e){if(e.key==='Escape')closePopup()})")
    f.write('</script></body></html>')

print('done')
