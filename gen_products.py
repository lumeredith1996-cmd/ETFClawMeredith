#!/usr/bin/env python3
"""Build all ETF site pages with popup, 2-channel tabs, single-select filters"""

import json
import openpyxl

# Load Longbridge quote data
try:
    lb_quotes = json.load(open('/root/.openclaw/workspace/etf-map-site/lb_quotes.json'))
    LB_QUOTES = lb_quotes.get('quotes', {})
except:
    LB_QUOTES = {}
from collections import defaultdict

# ===== READ PRODUCT LIST =====
wb = openpyxl.load_workbook('/root/.openclaw/media/inbound/all产品_20260419.xlsx')
ws = wb['products_all']
products = []
for row in ws.iter_rows(values_only=True):
    if row[0] is None or row[0] == '一级分类': continue
    products.append({
        'cat1': str(row[0]).strip(),
        'cat2': str(row[1]).strip(),
        'code': str(row[2]).strip() if row[2] else '',
        'name': str(row[3]).strip() if row[3] else '',
        'note': str(row[4]).strip() if row[4] else ''
    })

CROSSBORDER_CATS = {
    '跨境理财通（大湾区9市）-权益',
    '跨境理财通（大湾区9市）-货币',
    '跨境理财通（大湾区9市）-固定收益',
    '跨境理财通（大湾区9市）-基金互认',
}

# ===== BUILD CHANNEL membership =====
mainland_codes = set()
overseas_codes = set()
crossborder_codes = set()
for p in products:
    code = p['code']
    if p['cat1'] == '内地可买' and p['cat2'] not in CROSSBORDER_CATS:
        mainland_codes.add(code)
    if p['cat1'] == '内地可买' and p['cat2'] in CROSSBORDER_CATS:
        crossborder_codes.add(code)
    if p['cat1'] == '境外可买':
        overseas_codes.add(code)

code_channels = {}
for p in products:
    code = p['code']
    chs = code_channels.get(code, set())
    if code in mainland_codes: chs.add('mainland')
    if code in overseas_codes: chs.add('overseas')
    if code in crossborder_codes: chs.add('crossborder')
    code_channels[code] = chs
for p in products:
    p['channels'] = sorted(list(code_channels.get(p['code'], set())))

def merge_cat2(cat2):
    if cat2 == '个股杠反ETF-美元份额': return '个股杠反ETF'
    if cat2 == '货币市场ETF-美元份额': return '货币市场ETF'
    return cat2

CAT2_DISPLAY = {
    'QDII ETF': 'QDII ETF', 'QDII 联接基金': 'QDII 联接基金',
    '港股通ETF': '港股通ETF', '港股通ETF-60/40产品': '港股通60/40产品',
    '跨境理财通（大湾区9市）-权益': '跨境理财·权益',
    '跨境理财通（大湾区9市）-货币': '跨境理财·货币',
    '跨境理财通（大湾区9市）-固定收益': '跨境理财·固收',
    '跨境理财通（大湾区9市）-基金互认': '基金互认',
    '个股杠反ETF': '个股杠反ETF', '指数杠反ETF': '指数杠反ETF',
    '加密货币期货ETF': '加密期货ETF', '加密货币期货杠反ETF': '加密期货杠反',
    '商品期货杠反ETF': '商品期货杠反', '债券ETF-中国': '债券-中国',
    '债券ETF-美国': '债券-美国', '货币市场ETF': '货币市场ETF',
    'REITs ETF': 'REITs ETF', '香港': '香港股票ETF',
    '香港-美元份额': '香港-美元', '美国': '美国股票ETF',
    'A股': 'A股ETF', '新兴市场': '新兴市场', '日本': '日本股票ETF',
    '跨市场': '跨市场',
}
CAT2_COLOR = {
    'QDII ETF': '#3b82f6', 'QDII 联接基金': '#6366f1',
    '港股通ETF': '#10b981', '港股通ETF-60/40产品': '#14b8a6',
    '跨境理财通（大湾区9市）-权益': '#f59e0b',
    '跨境理财通（大湾区9市）-货币': '#fbbf24',
    '跨境理财通（大湾区9市）-固定收益': '#f97316',
    '跨境理财通（大湾区9市）-基金互认': '#eab308',
    '个股杠反ETF': '#ec4899', '指数杠反ETF': '#ef4444',
    '加密货币期货ETF': '#fbbf24', '加密货币期货杠反ETF': '#f59e0b',
    '商品期货杠反ETF': '#eab308', '债券ETF-中国': '#8b5cf6',
    '债券ETF-美国': '#6366f1', '货币市场ETF': '#10b981',
    'REITs ETF': '#14b8a6', '香港': '#ef4444',
    '香港-美元份额': '#dc2626', '美国': '#3b82f6',
    'A股': '#8b5cf6', '新兴市场': '#f59e0b', '日本': '#ec4899',
    '跨市场': '#14b8a6',
}

# ===== READ HOLDINGS DATA =====
def parse_holdings(path):
    try:
        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2['Performance Report']
        rows = list(ws2.iter_rows(values_only=True))
        data_date = rows[1][0].replace('数据截至日期: ', '') if len(rows) > 1 and rows[1][0] else ''
        holdings = []
        idx_section = None
        for r in rows[3:]:
            if r[0] == '主要市场股指':
                idx_section = True
                continue
            if r[0] in ('前二十大大持仓', '前二十大持仓', '前十大持仓') or (r[0] and '持仓' in str(r[0])):
                continue
            if idx_section and r[0] and r[1] is None:
                # market index row
                continue
            if r[1] and r[1] != '代码':
                name = str(r[0]) if r[0] else ''
                ticker = str(r[1]) if r[1] else ''
                region = str(r[2]) if r[2] else ''
                exchange = str(r[3]) if r[3] else ''
                try:
                    w1 = float(r[4]) if r[4] is not None else 0
                    m1 = float(r[5]) if r[5] is not None else 0
                    ytd = float(r[6]) if r[6] is not None else 0
                    wgt = float(r[9]) if r[9] is not None else 0
                except (ValueError, TypeError):
                    continue
                if ticker and name and wgt > 0:
                    holdings.append({'name': name, 'ticker': ticker, 'region': region,
                                     'exchange': exchange, 'w1': float(w1), 'm1': float(m1),
                                     'ytd': float(ytd), 'weight': float(wgt)})
        return data_date, holdings
    except Exception as e:
        print(f'Error reading {path}: {e}')
        return '', []

holdings_db = {}
holdings_files = {
    '513730.SH': '/root/.openclaw/media/inbound/ä_å_äº_ç_æ_ETF_2026-04-10---bb2540ac-8510-4621-a4f2-fecd4cf4ff72.xlsx',
    '159687.SZ': '/root/.openclaw/media/inbound/äº_å_ªç²¾é_ETF_2026-04-10---73e27b36-185e-4f17-8d69-39c201057957.xlsx',
    '3441.HK': '/root/.openclaw/media/inbound/å_æ_¹æ_è_å_æ_æ_è_è_ç_ç²¾é_ETF_2026-04-10---43206195-743b-4dd8-a066-d2bbaecc5d58.xlsx',
    '3442.HK': '/root/.openclaw/media/inbound/å_æ_¹æ_è_æ_ç_æ_ç¾_ç_æ_ETF_2026-04-10---02a799fe-55e9-45af-b620-288cd5b1b0aa.xlsx',
    '2830.HK': '/root/.openclaw/media/inbound/æ²_ç_¹é_æ_ä¼_ETF_2026-04-13---baa6fc66-98d6-428f-b8f7-6db42705cc21.xlsx',
}
for code, path in holdings_files.items():
    date, holdings = parse_holdings(path)
    if holdings:
        holdings_db[code] = {'date': date, 'holdings': holdings}
        print(f"Loaded {code}: {len(holdings)} holdings, date={date}")

# ===== HOLIDAY CALENDARS =====
def expand_date_range(ds, r):
    if '－' not in ds and '-' not in ds and '至' not in ds: return [(ds, r)]
    sep = '－' if '－' in ds else ('-' if '-' in ds else '至')
    p = ds.split(sep); s, e = p[0].strip(), p[1].strip() if len(p) > 1 else s
    import re
    ms = int(re.search(r'(\d+)月', s).group(1)); ds2 = int(re.search(r'(\d+)日', s).group(1))
    me = int(re.search(r'(\d+)月', e).group(1)); de = int(re.search(r'(\d+)日', e).group(1))
    if ms == me: return [(f'{ms}月{d}日', r) for d in range(ds2, de + 1)]
    return [(s, r), (e, r)]

def expand_holidays(raw_list):
    result = []
    for item in raw_list:
        result.extend(expand_date_range(item[0], item[1] if len(item) > 1 else ''))
    return result

HOLIDAYS_RAW = {
    '513730.SH': {
        'name': '东南亚科技ETF',
        'raw': [
            ('1月16日',), ('1月19日',), ('3月3日',),
            ('3月18日-3月20日',), ('3月23日-3月24日',),
            ('4月2日-4月3日',), ('4月13日-4月15日',),
            ('5月14日-5月15日',), ('5月25日-5月29日',),
            ('6月1日',), ('6月3日',), ('6月16日',),
            ('7月3日',), ('7月28日-7月29日',),
            ('8月7日',), ('8月10日',), ('8月12日',), ('8月17日',), ('8月25日',),
            ('9月7日',), ('10月13日',), ('10月23日',),
            ('11月6日',), ('11月9日',), ('11月26日',),
            ('12月7日',), ('12月10日',), ('12月24日',), ('12月25日',), ('12月31日',),
        ],
        'note': '泰国2026年休市日历'
    },
    '159687.SZ': {
        'name': '亚太精选ETF',
        'raw': [
            ('1月12日', '日本成人节'), ('1月19日', '美国马丁·路德·金节'),
            ('2月11日', '日本建国纪念日'),
            ('3月20日', '日本春分节'),
            ('4月3日', '香港/美国/新加坡耶稣受难日'), ('4月7日', '香港复活节假期'),
            ('4月29日', '日本昭和日'),
            ('5月6日', '日本宪法纪念日假期'),
            ('5月25日', '香港佛诞节/美国阵亡将士纪念日'), ('5月27日', '新加坡哈芝节'),
            ('6月1日', '新加坡卫塞节假期'),
            ('7月1日', '香港特区成立纪念日'), ('7月3日', '美国独立日假期'),
            ('7月20日', '日本海之日'),
            ('8月10日', '新加坡国庆日假期'), ('8月11日', '日本山之日'),
            ('9月7日', '美国劳动节'),
            ('9月21日', '日本敬老日'), ('9月22日',), ('9月23日', '日本秋分节假期'),
            ('10月12日', '日本体育日'), ('10月19日', '香港重阳节假期'),
            ('11月3日', '日本文化日'), ('11月9日', '新加坡屠妖节假期'),
            ('11月23日', '日本劳动感谢日'), ('11月26日', '美国感恩节'),
            ('12月24日', '港新圣诞节前日'), ('12月25日', '港美新圣诞节'),
            ('12月31日', '港新新年假期前日/日本银行假日'),
        ],
        'note': '亚太精选ETF 2026年休市日历（日本/美国/香港/新加坡）'
    },
}

# Build HOLIDAYS with expanded date ranges
HOLIDAYS = {}
for code in HOLIDAYS_RAW:
    HOLIDAYS[code] = {
        'name': HOLIDAYS_RAW[code]['name'],
        'holidays': expand_holidays(HOLIDAYS_RAW[code]['raw']),
        'note': HOLIDAYS_RAW[code]['note']
    }

# Link fund -> underlying ETF mapping
LINK_FUND_UNDERLYING = {
    '021189.OF': '159687.SZ',
    '020515.OF': '513730.SH',
    '020516.OF': '513730.SH',
}

# ===== PRODUCT DATABASE (for popup) =====
# Build a dict keyed by code -> product info
product_db = {}
for p in products:
    code = p['code']
    if code not in product_db:
        product_db[code] = {
            'name': p['name'], 'code': code, 'cat2': p['cat2'],
            'cat2_display': CAT2_DISPLAY.get(p['cat2'], p['cat2']),
            'cat2_color': CAT2_COLOR.get(p['cat2'], '#6b7280'),
            'note': p['note'],
            'channels': p['channels'],
            'holdings': holdings_db.get(code, None),
        }
        if code in HOLIDAYS:
            product_db[code]['holidays'] = HOLIDAYS[code]['holidays']
            product_db[code]['holidays_note'] = HOLIDAYS[code]['note']

# ===== INVESTMENT CHANNEL INFO =====
channel_info = {
    'QDII': {'icon': '🌐', 'label': 'QDII', 'region': '全国', 'threshold': '基本无门槛', 'quota': '每家基金管理人额度有限', 'target': '境外股票和债券'},
    '港股通ETF': {'icon': '🔗', 'label': '港股通ETF', 'region': '全国', 'threshold': '50万资产门槛', 'quota': '每日额度420亿港元', 'target': '港股（含ETF，不超过40%非港股仓位）'},
    '跨境理财通': {'icon': '🌏', 'label': '大湾区跨境理财通', 'region': '大湾区9市', 'threshold': '大湾区居民，有资产门槛', 'quota': '总额度1500亿人民币', 'target': '境外存款、债券、基金'},
    '境外直投': {'icon': '🏦', 'label': '境外账户直投', 'region': '全球', 'threshold': '需开境外账户', 'quota': '无额度限制', 'target': '全球各类型资产'},
}

def fmt_pct(v): return f'{v:+.2f}%' if v is not None else '—'

def product_popup_html(code, p, quote=None):
    # For link funds, use underlying ETF's holdings and holidays
    underlying = LINK_FUND_UNDERLYING.get(code)
    display_code = underlying if underlying else code
    h = holdings_db.get(display_code)  # Always look up from holdings_db for link funds
    if h:
        rows = ''
        for stock in h['holdings'][:10]:
            rows += f'''<tr>
  <td style="padding:6px 10px;font-size:12px">{stock['name']}</td>
  <td style="padding:6px 10px;font-size:11px;color:#3b82f6;font-family:monospace">{stock['ticker']}</td>
  <td style="padding:6px 10px;font-size:11px">{stock['region']}</td>
  <td style="padding:6px 10px;font-size:11px;text-align:right">{fmt_pct(stock['w1'])}</td>
  <td style="padding:6px 10px;font-size:11px;text-align:right">{fmt_pct(stock['m1'])}</td>
  <td style="padding:6px 10px;font-size:11px;text-align:right">{fmt_pct(stock['ytd'])}</td>
  <td style="padding:6px 10px;font-size:11px;text-align:right;font-weight:700;color:#10b981">{stock['weight']:.2f}%</td>
</tr>'''
        holdings_section = f'''<div style="margin-top:16px">
  <div style="font-size:12px;color:#9ca3af;margin-bottom:8px">数据截至: {h['date']} · 点击查看更多 →</div>
  <table style="width:100%;border-collapse:collapse">
    <thead>
      <tr style="background:#f0f2f8">
        <th style="padding:6px 10px;text-align:left;font-size:10px;color:#9ca3af">持仓名称</th>
        <th style="padding:6px 10px;text-align:left;font-size:10px;color:#9ca3af">代码</th>
        <th style="padding:6px 10px;text-align:left;font-size:10px;color:#9ca3af">地区</th>
        <th style="padding:6px 10px;text-align:right;font-size:10px;color:#9ca3af">1周</th>
        <th style="padding:6px 10px;text-align:right;font-size:10px;color:#9ca3af">1月</th>
        <th style="padding:6px 10px;text-align:right;font-size:10px;color:#9ca3af">YTD</th>
        <th style="padding:6px 10px;text-align:right;font-size:10px;color:#9ca3af">权重</th>
      </tr>
    </thead>
    <tbody>{rows}</tbody>
  </table>
</div>'''
    else:
        holdings_section = f'''<div style="margin-top:16px;padding:16px;background:#f8f9fc;border-radius:8px;text-align:center">
  <div style="font-size:24px;margin-bottom:8px">📋</div>
  <div style="font-size:13px;color:#6b7280">暂无成分股权重数据</div>
  <div style="font-size:12px;color:#9ca3af;margin-top:4px">如需补充请提供相关资料</div>
</div>'''

    # Holiday calendar - use underlying for link funds
    holidays_section = ''
    holiday_code = underlying if underlying else code
    holidays = HOLIDAYS.get(holiday_code, {}).get('holidays', [])
    holidays_note = HOLIDAYS.get(holiday_code, {}).get('note', '')
    if holidays:
        # Group by month
        by_month = defaultdict(list)
        for item in holidays:
            date_str = item[0]
            reason = item[1] if len(item) > 1 else ''
            month = date_str[:date_str.index('月')+1] if '月' in date_str else date_str
            by_month[month].append((date_str, reason))
        cal_rows = ''
        month_order = ['1月','2月','3月','4月','5月','6月','7月','8月','9月','10月','11月','12月']
        for month in month_order:
            if month not in by_month: continue
            items = by_month[month]
            day_cells = ''.join(f'<div style="padding:3px 6px;font-size:12px;border-radius:4px;background:#fff;border:1px solid #e2e5ef;min-width:80px"><div style="font-weight:600;color:#1a1d2e">{d}</div><div style="font-size:10px;color:#6b7280;margin-top:1px">{r}</div></div>' for d,r in items)
            cal_rows += f'<div style="margin-bottom:8px"><div style="font-size:11px;color:#9ca3af;margin-bottom:4px;font-weight:600">{month}</div><div style="display:flex;flex-wrap:wrap;gap:4px">{day_cells}</div></div>'
        holidays_section = f'''<div style="margin-top:16px;padding:14px;background:#f8f9ff;border-radius:10px;border:1px solid #e2e5ef">
          <div style="font-size:12px;font-weight:700;color:#1a1d2e;margin-bottom:10px">📅 2026年休市日历</div>
          <div style="font-size:11px;color:#6b7280;margin-bottom:10px">{holidays_note}</div>
          {cal_rows}
        </div>'''

    note_html = f'<div style="margin-top:10px;padding:8px 12px;background:#fffbeb;border-radius:6px;font-size:12px;color:#d97706">⚠️ {p["note"]}</div>' if p['note'] and p['note'] != 'None' else ''

    # Link fund extra info
    link_fund_extra = ''
    if underlying:
        link_fund_extra = f'<div style="margin-top:8px;padding:8px 12px;background:#eff6ff;border-radius:6px;font-size:12px;color:#1d4ed8"><strong>底层标的:</strong> {underlying}（持仓与休市日历同上）</div>'
    if underlying and p.get('note') and '限额' in p.get('note', ''):
        link_fund_extra += f'<div style="margin-top:6px;padding:6px 10px;background:#fef3c7;border-radius:6px;font-size:12px;color:#92400e"><strong>⚠️ 限额:</strong> {p["note"]}</div>'

    chan_labels = {'mainland':'🇨🇳 内地直接','overseas':'🌍 境外可买','crossborder':'🌐 跨境理财'}
    chan_html = ' '.join(f'<span style="background:#f0f2f8;padding:3px 8px;border-radius:4px;font-size:12px;margin-right:4px">{chan_labels.get(ch,ch)}</span>' for ch in p['channels'])

    # Quote section (Longbridge real-time data)
    quote_section = ''
    if quote:
        chg = quote.get('chg_pct', 0)
        chg_color = '#dc2626' if chg > 0 else ('#16a34a' if chg < 0 else '#6b7280')
        chg_arrow = '▲' if chg > 0 else ('▼' if chg < 0 else '—')
        quote_section = f'''<div style="margin-bottom:12px;padding:10px 14px;background:#f8f9fc;border-radius:8px;display:flex;align-items:center;justify-content:space-between">
          <div>
            <span style="font-size:22px;font-weight:700;font-family:monospace">{quote.get('last', '—')}</span>
            <span style="font-size:14px;margin-left:8px;color:{chg_color};font-weight:600">{chg_arrow} {chg:+.2f}%</span>
          </div>
          <div style="text-align:right;font-size:11px;color:#6b7280">
            <div>成交量: {quote.get('volume', '—')}</div>
            <div>昨收: {quote.get('prev_close', '—')}</div>
          </div>
        </div>'''
    elif code in LB_QUOTES:
        q = LB_QUOTES[code]
        chg = q.get('chg_pct', 0)
        chg_color = '#dc2626' if chg > 0 else ('#16a34a' if chg < 0 else '#6b7280')
        chg_arrow = '▲' if chg > 0 else ('▼' if chg < 0 else '—')
        quote_section = f'''<div style="margin-bottom:12px;padding:10px 14px;background:#f8f9fc;border-radius:8px;display:flex;align-items:center;justify-content:space-between">
          <div>
            <span style="font-size:22px;font-weight:700;font-family:monospace">{q.get('last','—')}</span>
            <span style="font-size:14px;margin-left:8px;color:{chg_color};font-weight:600">{chg_arrow} {chg:+.2f}%</span>
          </div>
          <div style="text-align:right;font-size:11px;color:#6b7280">
            <div>成交量: {q.get('volume','—')}</div>
            <div>昨收: {q.get('prev_close','—')}</div>
          </div>
        </div>'''

        return f'''<div class="popup-content">
  <div style="display:flex;align-items:flex-start;justify-content:space-between;gap:12px;margin-bottom:12px">
    <div>
      <div style="font-size:18px;font-weight:700;color:#1a1d2e">{p['name']}</div>
      <div style="font-size:13px;color:#3b82f6;margin-top:2px;font-family:monospace">{code}</div>
    </div>
    <span style="background:{p['cat2_color']}18;color:{p['cat2_color']};border:1px solid {p['cat2_color']}33;border-radius:4px;padding:4px 10px;font-size:12px;white-space:nowrap">{p['cat2_display']}</span>
  </div>
  <div style="margin-bottom:12px">{chan_html}</div>{note_html}{link_fund_extra}
  {holdings_section}
  {holidays_section}
</div>'''

# ===== BUILD PRODUCT ROWS =====
def esc(s):
    return s.replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;').replace("'",'&#39;')

def badge_html(cat2):
    d = CAT2_DISPLAY.get(cat2, cat2)
    c = CAT2_COLOR.get(cat2, '#6b7280')
    return f'<span style="background:{c}18;color:{c};border:1px solid {c}33;border-radius:4px;padding:2px 8px;font-size:12px;white-space:nowrap;display:inline-block">{d}</span>'

rows_html = ''
popup_data = {}

for p in products:
    code = p['code']
    code_disp = esc(code) if code else '—'
    name_disp = esc(p['name'])
    merged_cat2 = merge_cat2(p['cat2'])
    channels_str = ','.join(p['channels'])
    popup_data[code] = product_popup_html(code, product_db.get(code, p), LB_QUOTES.get(code))
    note_html = f'<span style="color:#d97706;font-size:11px;margin-left:6px">⚠️</span>' if p['note'] and p['note'] != 'None' else ''
    rows_html += f'''<tr class="product-row" data-channels="{channels_str}" data-cat2="{esc(merged_cat2)}" data-code="{esc(code)}" data-name="{name_disp}" onclick="showPopup('{esc(code)}')" style="cursor:pointer">
  <td style="padding:10px 16px"><code style="color:#3b82f6;font-size:13px">{code_disp}</code></td>
  <td style="padding:10px 16px"><strong>{name_disp}</strong>{note_html}</td>
  <td style="padding:10px 16px">{badge_html(p['cat2'])}</td>
</tr>'''

# ===== CHANNEL COUNTS =====
cnt_mainland = len(set(p['code'] for p in products if 'mainland' in p['channels']))
cnt_overseas = len(set(p['code'] for p in products if 'overseas' in p['channels']))

# ===== SUB-CATEGORY BUTTONS =====
def get_cat2_set(channel):
    cats = set()
    for p in products:
        if channel in p['channels']:
            cats.add(merge_cat2(p['cat2']))
    return sorted(cats)

def make_cat2_buttons(channel):
    cats = get_cat2_set(channel)
    btns = ''
    for c in cats:
        d = CAT2_DISPLAY.get(c, c)
        c_esc = esc(c)
        btns += f'<button class="sub-btn" data-cat2="{c_esc}" onclick="toggleCat2(\'{c_esc}\', this)">{d}</button>'
    return btns

mainland_btns = make_cat2_buttons('mainland')
overseas_btns = make_cat2_buttons('overseas')

# ===== INVESTMENT CHANNEL TABLE =====
channel_rows = ''
for ch_id, info in channel_info.items():
    channel_rows += f'''<tr>
  <td style="padding:14px 16px;border-bottom:1px solid #f0f2f8">
    <span style="font-size:18px;margin-right:8px">{info['icon']}</span>
    <strong style="font-size:14px">{info['label']}</strong>
  </td>
  <td style="padding:14px 16px;border-bottom:1px solid #f0f2f8;font-size:13px">{info['region']}</td>
  <td style="padding:14px 16px;border-bottom:1px solid #f0f2f8;font-size:12px;color:#6b7280">{info['threshold']}</td>
  <td style="padding:14px 16px;border-bottom:1px solid #f0f2f8;font-size:12px;color:#6b7280">{info['quota']}</td>
  <td style="padding:14px 16px;border-bottom:1px solid #f0f2f8;font-size:13px">{info['target']}</td>
</tr>'''

# ===== JS POPUP DATA =====
popup_json = json.dumps({code: popup_data[code] for code in popup_data}, ensure_ascii=False)

print(f"Products: {len(products)}, Popup codes: {len(popup_data)}")
print(f"Mainland: {cnt_mainland}, Overseas: {cnt_overseas}")

html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>产品筛选 — 全球ETF投资工具箱</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f5f6fa;color:#1a1d2e;min-height:100vh}}
.header{{background:#1a1d2e;color:#fff;padding:18px 24px;display:flex;align-items:center;gap:16px;flex-wrap:wrap}}
.header h1{{font-size:20px;font-weight:700}}
.header p{{font-size:13px;opacity:0.65}}
.nav{{display:flex;gap:8px;padding:12px 24px;background:#fff;border-bottom:1px solid #e2e5ef;flex-wrap:wrap}}
.nav a{{color:#6b7280;text-decoration:none;font-size:14px;padding:6px 14px;border-radius:6px;transition:all 0.15s}}
.nav a:hover{{background:#f0f2f8;color:#1a1d2e}}
.nav a.active{{background:#1a1d2e;color:#fff}}

.channel-bar{{padding:20px 24px 0;background:#fff;border-bottom:2px solid #e2e5ef}}
.channel-label{{font-size:11px;color:#9ca3af;font-weight:700;letter-spacing:0.1em;text-transform:uppercase;margin-bottom:10px}}
.channel-tabs{{display:flex;gap:8px;flex-wrap:wrap}}
.chan-btn{{flex:1;min-width:160px;max-width:260px;padding:12px 16px;border:2px solid #e2e5ef;border-bottom:none;border-radius:10px 10px 0 0;font-size:13px;font-weight:700;cursor:pointer;background:#f8f9fc;color:#6b7280;transition:all 0.15s;text-align:center}}
.chan-btn:hover{{background:#f0f2f8;color:#1a1d2e}}
.chan-btn.active{{background:#1a1d2e;color:#fff;border-color:#1a1d2e}}

.sub-bar{{padding:12px 24px;background:#fff;display:flex;gap:8px;flex-wrap:wrap;align-items:center;border-bottom:1px solid #f0f2f8;min-height:52px}}
.sub-bar-label{{font-size:12px;color:#9ca3af;font-weight:600;margin-right:4px;flex-shrink:0}}
.sub-btn{{padding:6px 14px;border:1px solid #e2e5ef;border-radius:20px;font-size:13px;cursor:pointer;background:#fff;color:#6b7280;transition:all 0.15s;flex-shrink:0}}
.sub-btn:hover{{border-color:#3b82f6;color:#3b82f6}}
.sub-btn.active{{background:#1a1d2e;color:#fff;border-color:#1a1d2e}}

.search-bar{{padding:12px 24px;background:#fff;border-bottom:1px solid #f0f2f8;display:flex;gap:12px;align-items:center}}
.search-input{{flex:1;padding:10px 14px;border:1px solid #e2e5ef;border-radius:8px;font-size:14px;outline:none;max-width:400px;transition:border 0.15s}}
.search-input:focus{{border-color:#3b82f6}}

.stats{{padding:8px 24px;font-size:12px;color:#6b7280;background:#f8f9fc}}

table{{width:100%;border-collapse:collapse}}
thead{{background:#f0f2f8;position:sticky;top:0;z-index:10}}
th{{text-align:left;padding:10px 16px;font-size:11px;font-weight:600;color:#9ca3af;text-transform:uppercase;letter-spacing:0.06em;border-bottom:2px solid #e2e5ef}}
td{{padding:10px 16px;font-size:14px;border-bottom:1px solid #f0f2f8;vertical-align:middle}}
tr.product-row:hover{{background:#f8fbff}}
.product-row{{cursor:pointer}}

.no-result{{text-align:center;padding:60px;color:#9ca3af}}
.no-result div{{font-size:40px;margin-bottom:12px}}

/* Channel compare table */
#channelPanel{{display:none}}
#channelPanel.active{{display:block}}
.ch-table{{width:100%;border-collapse:collapse;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,0.06);border:1px solid #e2e5ef}}
.ch-table thead th{{background:#1a1d2e;color:#fff;font-size:11px;padding:10px 16px;text-align:left}}
.ch-table td{{font-size:13px}}
.ch-table tr:hover{{background:#f8fbff}}
.ch-intro{{padding:16px 24px;background:#fff;border-bottom:1px solid #f0f2f8;font-size:13px;color:#6b7280;line-height:1.6}}

/* List panel */
#listPanel{{display:block}}
#listPanel.hidden{{display:none}}

/* Popup */
#popup{{display:none;position:fixed;inset:0;z-index:1000;background:rgba(0,0,0,0.5);backdrop-filter:blur(4px);justify-content:center;align-items:center;padding:20px}}
#popup.show{{display:flex}}
.popup-box{{background:#fff;border-radius:16px;max-width:720px;width:100%;max-height:85vh;overflow-y:auto;box-shadow:0 20px 60px rgba(0,0,0,0.3);position:relative}}
.popup-header{{padding:16px 20px;border-bottom:1px solid #f0f2f8;display:flex;align-items:center;justify-content:space-between}}
.popup-header h3{{font-size:16px;font-weight:700}}
.popup-close{{width:32px;height:32px;border-radius:50%;border:none;background:#f0f2f8;cursor:pointer;font-size:18px;display:flex;align-items:center;justify-content:center;color:#6b7280;transition:all 0.15s;flex-shrink:0}}
.popup-close:hover{{background:#e2e5ef;color:#1a1d2e}}
.popup-body{{padding:20px}}
</style>
</head>
<body>
<div class="header">
  <div>
    <h1>🔎 产品筛选</h1>
    <p>共 {len(products)} 只ETF产品 · 点击行查看持仓明细+休市日历+实时行情</p>
  </div>
</div>
<nav class="nav">
  <a href="index.html">🗺️ 全球ETF地图</a>
  <a href="products.html" class="active">🔎 产品筛选</a>
</nav>

<div class="channel-bar">
  <div class="channel-label">投资渠道</div>
  <div class="channel-tabs">
    <button class="chan-btn active" id="chan-mainland" onclick="switchChannel('mainland')">
      🇹🇨 内地可买 <span style="opacity:0.65;font-size:11px;font-weight:400">({cnt_mainland})</span>
    </button>
    <button class="chan-btn" id="chan-overseas" onclick="switchChannel('overseas')">
      🌍 境外可买 <span style="opacity:0.65;font-size:11px;font-weight:400">({cnt_overseas})</span>
    </button>
  </div>
</div>

<!-- 列表视图 -->
<div id="listPanel">
  <div class="sub-bar" id="subBar">
    <span class="sub-bar-label">类型筛选:</span>
    <div id="subBtns">{mainland_btns}</div>
  </div>
  <div class="search-bar">
    <input type="text" class="search-input" id="searchInput" placeholder="🔍  搜索产品名称或代码..." oninput="applyFilters()">
  </div>
  <div class="stats" id="statsBar">显示全部 {cnt_mainland} 只产品</div>
  <div style="overflow-x:auto">
    <table>
    <thead>
      <tr>
        <th style="width:120px">代码</th>
        <th>产品名称</th>
        <th style="width:180px">类型</th>
      </tr>
    </thead>
    <tbody id="tableBody">
    {rows_html}
    </tbody>
    </table>
  </div>
  <div id="noResult" class="no-result" style="display:none"><div>🔍</div><div>没有找到匹配的产品</div></div>
</div>

<!-- 渠道对比视图 -->
<div id="channelPanel">
  <div class="ch-intro">
    <strong>四大跨境投资渠道对比：</strong>不同渠道在投资门槛、额度限制、可投标的范围等方面各有不同，适合不同需求的投资者。
  </div>
  <div style="padding:20px 24px;overflow-x:auto">
    <table class="ch-table">
    <thead>
      <tr>
        <th style="width:150px">渠道</th>
        <th style="width:80px">可参与地区</th>
        <th style="width:180px">投资门槛</th>
        <th style="width:160px">额度限制</th>
        <th>资金投向</th>
      </tr>
    </thead>
    <tbody>
    {channel_rows}
    </tbody>
    </table>
  </div>
</div>

<!-- 弹窗 -->
<div id="popup" onclick="closePopupOnBg(event)">
  <div class="popup-box">
    <div class="popup-header">
      <h3>📋 产品详情</h3>
      <button class="popup-close" onclick="closePopup()">×</button>
    </div>
    <div class="popup-body" id="popupBody"></div>
  </div>
</div>

<script>
const POPUP_DATA = {popup_json};
const CHANNEL_SUBBTNS = {{
  mainland: `{mainland_btns}`,
  overseas: `{overseas_btns}`
}};
const CHANNEL_COUNTS = {{mainland: {cnt_mainland}, overseas: {cnt_overseas}}};


let curChannel = 'mainland';
let activeCat2 = null; // 单选
const rows = document.querySelectorAll('#tableBody tr');
const searchInput = document.getElementById('searchInput');
const statsBar = document.getElementById('statsBar');
const noResult = document.getElementById('noResult');
const subBtns = document.getElementById('subBtns');
const listPanel = document.getElementById('listPanel');
const channelPanel = document.getElementById('channelPanel');

function switchChannel(ch) {{
  curChannel = ch;
  activeCat2 = null;
  document.querySelectorAll('.chan-btn').forEach(b => b.classList.remove('active'));
  document.getElementById('chan-' + ch).classList.add('active');
  listPanel.classList.remove('hidden');
  channelPanel.classList.remove('active');
  subBtns.innerHTML = '<span class="sub-bar-label">类型筛选:</span>' + CHANNEL_SUBBTNS[ch];
  applyFilters();
}}

function toggleCat2(cat2, btn) {{
  // 单选逻辑
  if (activeCat2 === cat2) {{
    activeCat2 = null;
    btn.classList.remove('active');
  }} else {{
    document.querySelectorAll('.sub-btn').forEach(b => b.classList.remove('active'));
    activeCat2 = cat2;
    btn.classList.add('active');
  }}
  applyFilters();
}}

function applyFilters() {{
  const q = searchInput.value.trim().toLowerCase();
  let visible = 0;
  rows.forEach(row => {{
    const channels = row.dataset.channels.split(',');
    const cat2 = row.dataset.cat2;
    const code = row.dataset.code.toLowerCase();
    const name = row.dataset.name.toLowerCase();
    const chanOk = channels.includes(curChannel);
    const cat2Ok = !activeCat2 || cat2 === activeCat2;
    const searchOk = !q || code.includes(q) || name.includes(q);
    const show = chanOk && cat2Ok && searchOk;
    row.style.display = show ? '' : 'none';
    if (show) visible++;
  }});
  statsBar.textContent = `显示 ${{visible}} / ${{CHANNEL_COUNTS[curChannel]}} 只产品`;
  noResult.style.display = visible === 0 ? 'block' : 'none';
}}

function showPopup(code) {{
  const data = POPUP_DATA[code];
  if (!data) return;
  document.getElementById('popupBody').innerHTML = data;
  document.getElementById('popup').classList.add('show');
  document.body.style.overflow = 'hidden';
}}

function closePopup() {{
  document.getElementById('popup').classList.remove('show');
  document.body.style.overflow = '';
}}

function closePopupOnBg(e) {{
  if (e.target === document.getElementById('popup')) closePopup();
}}

document.addEventListener('keydown', e => {{ if (e.key === 'Escape') closePopup(); }});

applyFilters();
</script>
</body>
</html>'''

with open('/root/.openclaw/workspace/etf-map-site/products.html', 'w', encoding='utf-8') as f:
    f.write(html)
print("✅ products.html 生成完成")
